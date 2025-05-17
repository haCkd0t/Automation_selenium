import time

from selenium.webdriver.common.by import By
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains

workbook = openpyxl.load_workbook("C:\\Users\\joke\\Downloads\\file_example_XLS_10 (1).xlsx")
sheet = workbook['Sheet1']
max_row = sheet.max_row
max_col = sheet.max_column

driver = webdriver.Chrome()
driver.get("https://demoqa.com/webtables")
for r in range(1,max_row+1):
    driver.find_element(By.XPATH,"//button[@id='addNewRecordButton']").click()
    first_name = sheet.cell(r, 1).value
    last_name = sheet.cell(r, 2).value
    email = sheet.cell(r, 3).value
    age = sheet.cell(r, 4).value
    sal = sheet.cell(r, 5).value
    dep = sheet.cell(r, 6).value
    driver.find_element(By.XPATH,value="//input[@placeholder='First Name']").send_keys(first_name)
    driver.find_element(By.XPATH, value="//input[@placeholder='Last Name']").send_keys(last_name)
    driver.find_element(By.XPATH, value="//input[@placeholder='name@example.com']").send_keys(email)
    driver.find_element(By.XPATH, value="//input[@placeholder='Salary']").send_keys(sal)
    driver.find_element(By.XPATH, value="//input[@placeholder='Age']").send_keys(age)
    driver.find_element(By.XPATH, value="//input[@placeholder='Department']").send_keys(dep)
    driver.find_element(By.XPATH,value="//button[@id='submit']").click()
    time.sleep(2)

